import os, torch, re, sys, json, zipfile, json, shutil, traceback
from pprint import pprint
import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from pandas import ExcelWriter
from openpyxl.utils import get_column_letter


def parse_env_file(env_path):
    env = {}
    try:
        with open(env_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                if line.startswith("export "):
                    line = line[len("export "):]
                if "=" in line:
                    key, val = line.split("=", 1)
                    key = key.strip()
                    val = val.strip().strip('"').strip("'")
                    env[key] = val
    except FileNotFoundError:
        pass
    return env


def get_session_name(env_path):
    env = parse_env_file(env_path)
    session = env.get("session_name") or env.get("SESSION_NAME")
    if not session:
        raise RuntimeError(f"Missing 'session_name' in .env at {env_path}")
    return session


def list_run_dirs(runs_root):
    if not os.path.isdir(runs_root):
        raise RuntimeError(f"runs_data directory not found: {runs_root}")
    run_dirs = []
    for name in os.listdir(runs_root):
        p = os.path.join(runs_root, name)
        if os.path.isdir(p):
            run_dirs.append(p)
    run_dirs.sort()
    return run_dirs


def collect_target_files(run_dir):
    targets = []
    runsDataDf = []
    for root, _, files in os.walk(run_dir):
        for fn in files:
            ext = os.path.splitext(fn)[1].lower()
            if ext in (".png", ".json"):
                targets.append(os.path.join(root, fn))

            # Check for .pth files as well
            if ext in [".pth"]:
                if "Backups" not in root:
                    try:
                        with open(os.path.join(root, "conf.json"), 'r', encoding='utf-8') as file:
                            config = json.load(file)

                        trainingData = torch.load(os.path.join(root, fn), weights_only = False)
                        saveDf = pd.DataFrame(trainingData.get("train_history"))
                        dup_counts = saveDf['seed'].value_counts() - 1
                        dup_counts = dup_counts[dup_counts > 0]

                        _duplicateSeedCount = dup_counts.sum()
                        _episodeCount = saveDf['episode'].max()
                        _maxTimesteps = saveDf['timesteps'].max()
                        _minTimesteps = saveDf['timesteps'].min()
                        _avgTimesteps = saveDf['timesteps'].mean()
                        _TotalDuration = saveDf['duration'].sum()
                        _maxEpisodeDuration = saveDf['duration'].max()
                        _avgEpisodeDuration = saveDf['duration'].mean()
                        _minEpisodeDuration = saveDf['duration'].min()
                        _totalTimesteps = saveDf['timesteps'].sum()
                        _maxPoints = saveDf['points'].max()
                        
                        # Add data to the dataframe. Do not add the backup folder
                        if "Backups" not in root:
                            # Adapt for legacy training runs. TODO: Irrelevant in publication of main repo
                            if not ("algorithm" in config):
                                if ("decay" in config) and ("gamma" in config) and ("batch" in config):
                                    config["algorithm"] = "DDQN"

                                    if not ("algorithm_options" in config):
                                        config["algorithm_options"] = {
                                            "decay": config["decay"] if "decay" in config else -1,
                                            "batch": config["batch"] if "batch" in config else -1,
                                            "gamma": config["gamma"] if "gamma" in config else -1,
                                            "memorySize": config["memorySize"] if "memorySize" in config else -1,
                                            "startEbsilon": config["startEbsilon"] if "startEbsilon" in config else -1,
                                            "endEbsilon": config["endEbsilon"] if "endEbsilon" in config else -1,
                                            "numUpdateTS": config["numUpdateTS"] if "numUpdateTS" in config else -1,
                                        }
                                
                                if ("clip" in config) and ("nUpdatesPerIteration" in config) :
                                    config["algorithm"] = "DDQN"

                                    if not ("algorithm_options" in config):
                                        config["algorithm_options"] = {
                                            "gamma": config["gamma"] if "gamma" in config else -1,
                                            "clip": config["clip"] if "clip" in config else -1,
                                            "nUpdatesPerIteration": config["nUpdatesPerIteration"] if "nUpdatesPerIteration" in config else -1,
                                            "timeStepsPerBatch": config["timeStepsPerBatch"] if "timeStepsPerBatch" in config else -1,
                                            "entropyCoef": config["entropyCoef"] if "entropyCoef" in config else -1,
                                            "advantage_method": config["advantage_method"] if "advantage_method" in config else -1,
                                            "gae_lambda": config["gae_lambda"] if "gae_lambda" in config else -1,
                                            "maxNumTimeSteps": config["maxNumTimeSteps"] if "maxNumTimeSteps" in config else -1,
                                        }
                                
                            if not("env" in config):
                                config["env"] = "LunarLander-v3"

                            if ("algorithm_options" in config) and  not ("learning_rate" in config["algorithm_options"]):
                                if "learning_rate" in config: 
                                    config["algorithm_options"]["learning_rate"] = config["learning_rate"]
                                else:
                                    raise Exception("ERR!")

                            if ("architecture" in config) and (config["architecture"] == "snn"):
                                
                                if config["architecture"] == "snn":
                                    if config["algorithm"] == "PPO":
                                        config["network_actor"] = config["architecture"]
                                        config["network_actor_options"] = {
                                            "hidden_layers": config["hidden_layers"],
                                            "snn_tSteps": config["snn_tSteps"],
                                            "snn_beta": config["snn_beta"]
                                        }

                                        config["network_critic"] = config["architecture"]
                                        config["network_critic_options"] = {
                                            "hidden_layers": config["hidden_layers"],
                                            "snn_tSteps": config["snn_tSteps"],
                                            "snn_beta": config["snn_beta"]
                                        }
                                    if config["algorithm"] == "DDQN":
                                        config["network"] = config["architecture"]
                                        config["network_options"] = {
                                            "hidden_layers": config["hidden_layers"],
                                            "snn_tSteps": config["snn_tSteps"],
                                            "snn_beta": config["snn_beta"]
                                        }

                                if config["architecture"] == "ann":
                                    if config["algorithm"] == "PPO":
                                        config["network_actor"] = config["architecture"]
                                        config["network_actor_options"] = {
                                            "hidden_layers": config["hidden_layers"],
                                        }

                                        config["network_critic"] = config["architecture"]
                                        config["network_critic_options"] = {
                                            "hidden_layers": config["hidden_layers"],
                                        }
                                    if config["algorithm"] == "DDQN":
                                        config["network"] = config["architecture"]
                                        config["network_options"] = {
                                            "hidden_layers": config["hidden_layers"],
                                        }
                            
                            if ("network_options" in config) and ("hidden_layers_actor" in config["network_options"]):
                                if config["network"] == "ann":
                                    config["network_actor"] = config["network"]
                                    config["network_actor_options"] = {
                                        "hidden_layers": config["network_options"]["hidden_layers_actor"],
                                    }

                                    config["network_critic"] = config["network"]
                                    config["network_critic_options"] = {
                                        "hidden_layers": config["network_options"]["hidden_layers_critic"],
                                    }
                                elif config["network"] == "snn":
                                    config["network_actor"] = config["network"]
                                    config["network_actor_options"] = {
                                        "hidden_layers": config["network_options"]["hidden_layers_actor"],
                                        "snn_tSteps": config["network_options"]["snn_tSteps"],
                                        "snn_beta": config["network_options"]["snn_beta"],
                                    }

                                    config["network_critic"] = config["network"]
                                    config["network_critic_options"] = {
                                        "hidden_layers": config["network_options"]["hidden_layers_critic"],
                                        "snn_tSteps": config["network_options"]["snn_tSteps"],
                                        "snn_beta": config["network_options"]["snn_beta"],
                                    }

                            _dict = {
                                "session": os.getenv("session_name"),
                                "run": root.split(os.sep)[-1], # Run's number
                                "episodes": _episodeCount,
                                "numDuplicateInitialConditions": _duplicateSeedCount,
                                "trainingDuration": _TotalDuration,
                                "maxEpisodeDuration": _maxEpisodeDuration,
                                "avgEpisodeDuration": _avgEpisodeDuration,
                                "minEpisodeDuration": _minEpisodeDuration,
                                "totalTrainingTimesteps": _totalTimesteps,
                                "maxEpisodeTimesteps": _maxTimesteps,
                                "avgEpisodeTimesteps": _avgTimesteps,
                                "minEpisodeTimesteps": _minTimesteps,
                                "maxPoint": _maxPoints,
                                "algorithm": config["algorithm"],
                                "environment": config["env"] if "env" in config else "*",
                                "learning_rate": config["algorithm_options"]["learning_rate"],
                                "maxNumTimeSteps": config["algorithm_options"]["maxNumTimeSteps"] if "maxNumTimeSteps" in config["algorithm_options"] else -1 ,
                                
                                # PPO algogrithm
                                "gamma": config["algorithm_options"]["gamma"] if config["algorithm"] == "PPO" else "*",
                                "clip": config["algorithm_options"]["clip"] if config["algorithm"] == "PPO" else "*",
                                "nUpdatesPerIteration": config["algorithm_options"]["nUpdatesPerIteration"] if config["algorithm"] == "PPO" else "*",
                                "timeStepsPerBatch": config["algorithm_options"]["timeStepsPerBatch"] if config["algorithm"] == "PPO" else "*",
                                "entropyCoef": config["algorithm_options"]["entropyCoef"] if config["algorithm"] == "PPO" else "*",
                                "advantage_method": config["algorithm_options"]["advantage_method"] if config["algorithm"] == "PPO" else "*",
                                "gae_lambda": config["algorithm_options"]["gae_lambda"] if config["algorithm"] == "PPO" else "*",
                                
                                # DDQN algorithm
                                "decay": config["algorithm_options"]["decay"] if config["algorithm"] == "DDQN" else "*",
                                "batch": config["algorithm_options"]["batch"] if config["algorithm"] == "DDQN" else "*",
                                "gamma": config["algorithm_options"]["gamma"] if config["algorithm"] == "DDQN" else "*",
                                "memorySize": config["algorithm_options"]["memorySize"] if config["algorithm"] == "DDQN" else "*",
                                "startEbsilon": config["algorithm_options"]["startEbsilon"] if config["algorithm"] == "DDQN" else "*",
                                "endEbsilon": config["algorithm_options"]["endEbsilon"] if config["algorithm"] == "DDQN" else "*",
                                "numUpdateTS": config["algorithm_options"]["numUpdateTS"] if config["algorithm"] == "DDQN" else "*",

                                # Network details
                                "network_1_type": config["network_actor"] if config["algorithm"] == "PPO" else config["network"],
                                "network_1_details": config["network_actor_options"]["snn_tSteps"] if config["algorithm"] == "PPO" and "snn_tSteps" in config["network_critic_options"] == "PPO" else "*",
                                "network_1_layers": config["network_actor_options"]["hidden_layers"] if config["algorithm"] == "PPO" else config["network_options"]["hidden_layers"],
                                "network_2_type": config["network_critic"] if config["algorithm"] == "PPO" else config["network"],
                                "network_2_details": config["network_critic_options"]["snn_tSteps"] if config["algorithm"] == "PPO" and "snn_tSteps" in config["network_critic_options"] == "PPO" else "*",
                                "network_2_layers": config["network_critic_options"]["hidden_layers"] if config["algorithm"] == "PPO" else config["network_options"]["hidden_layers"],
                            }
                            runsDataDf.append(_dict)

                    except Exception as e:
                        # from pprint import pprint
                        # print("aaaaaa", e)
                        # print("file:", root)
                        # pprint(config)
                        # traceback.print_exc()   
                        # exit()

                        # Add data to the dataframe
                        runsDataDf.append({
                            "session": os.getenv("session_name"),
                            "run": root.split(os.sep)[-1], # Run's number
                            "episodes": -1,
                            "numDuplicateInitialConditions": -1,
                            "trainingDuration": -1,
                            "maxEpisodeDuration": -1,
                            "avgEpisodeDuration": -1,
                            "minEpisodeDuration": -1,
                            "totalTrainingTimesteps": -1,
                            "maxEpisodeTimesteps": -1,
                            "avgEpisodeTimesteps": -1,
                            "minEpisodeTimesteps": -1,
                            "maxPoint": -1,
                        })


    return targets, runsDataDf


def derive_run_folder_name(run_dir_name):
    m = re.search(r"\d+", run_dir_name)
    return m.group(0) if m else run_dir_name


def create_zip(session_name, runs_root, output_zip):
    run_dirs = list_run_dirs(runs_root)
    included_files = 0

    # Save the details of runs into an excel file
    runsDataDf = []
    for run_dir in run_dirs:
        _, tmp = collect_target_files(run_dir)
        runsDataDf = runsDataDf + tmp
    
    # Save the dataframe to excel
    with ExcelWriter(f"runs_data/{session_name}_trainingRunsData.xlsx", engine="openpyxl") as writer:
        runsDataDf = pd.DataFrame(runsDataDf)
        runsDataDf.to_excel(writer, index=False, sheet_name="Sheet1")

        # Get the workbook and sheet
        workbook = writer.book
        sheet = workbook["Sheet1"]

        # Get necessary column numbers
        headerMap = {cell.value: cell.column for cell in sheet[1]}
        ppoStartColNum = headerMap.get("gamma")
        ppoEndColNum = headerMap.get("gae_lambda")
        ddqnStartColNum = headerMap.get("decay")
        ddqnEndColNum = headerMap.get("numUpdateTS")

        # Make header for each algorithm
        sheet.insert_rows(1)

        for i in range(1,ppoStartColNum):
            sheet.cell(row=1, column = i).value = sheet.cell(row=2, column = i).value
            sheet.merge_cells(start_row =  1, start_column = i, end_row = 2, end_column = i)
        for i in range(ddqnEndColNum+1,sheet.max_column+1):
            sheet.cell(row=1, column = i).value = sheet.cell(row=2, column = i).value
            sheet.merge_cells(start_row =  1, start_column = i, end_row = 2, end_column = i)
        
        # Merge cells for PPO and DDQN headers
        # PPO starts at column ppoStartColNum and ends  on ppoEndColNum
        sheet.merge_cells(start_row =  1, start_column = ppoStartColNum, end_row = 1, end_column = ppoEndColNum)
        sheet.cell(row=1, column = ppoStartColNum).value = "PPO"
        
        # DDQN starts at column ddqnStartColNum and ends on ddqnEndColNum
        sheet.merge_cells(start_row =  1, start_column = ddqnStartColNum, end_row = 1, end_column = ddqnEndColNum)
        sheet.cell(row=1, column = ddqnStartColNum).value = "DDQN"

        # Style cells
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = Font(name="Calibri", bold=False)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set cell widths
        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                try:
                    cell_len = len(str(cell.value))
                    if cell_len > max_length:
                        max_length = cell_len
                except:
                    pass

            # Add small extra space for padding
            sheet.column_dimensions[col_letter].width = max_length + .2

    
    with zipfile.ZipFile(output_zip, "w", compression=zipfile.ZIP_DEFLATED) as zf:

        for run_dir in run_dirs:
            run_dir_name = os.path.basename(run_dir)
            run_folder_in_zip = derive_run_folder_name(run_dir_name)
            files, _ = collect_target_files(run_dir)

            for fp in files:
                rel = os.path.relpath(fp, start=run_dir)
                arc = os.path.join(run_folder_in_zip, rel).replace("\\", "/")
                zf.write(fp, arcname=arc)
                included_files += 1

        # Adding the excel file
        if os.path.exists(f"runs_data/{session_name}_trainingRunsData.xlsx"):
            zf.write(f"runs_data/{session_name}_trainingRunsData.xlsx", arcname=f"{session_name}_trainingRunsData.xlsx")
            print(f"Added runs_data/{session_name}_trainingRunsData.xlsx as {session_name}_trainingRunsData.xlsx")
        else:
            print(f"Warning: runs_data/{session_name}_trainingRunsData.xlsx not found, skipping.")

    return len(run_dirs), included_files


def get_telegram_credentials(env_path):
    env = parse_env_file(env_path)
    chat_id = env.get("telegram_chat_id")
    bot_token = env.get("telegram_bot_token")
    if not chat_id or not bot_token:
        raise RuntimeError("Missing 'telegram_chat_id' or 'telegram_bot_token' in .env")
    return chat_id, bot_token


def getMatchingRunConfigs(runs_dir="runs_data", current_algorithm=None):
    """
    Finds all run configurations in the specified directory that match the current algorithm.

    Args:
        runs_dir (str): The directory to search for run configurations. Defaults to "runs_data".
        current_algorithm (str, optional): The current algorithm to match. If None, attempts to read from "conf.json".

    Returns:
        list: A list of matching run configurations.
    """
    if current_algorithm is None:
        root_conf_path = os.path.join(".", "conf.json")
        if os.path.isfile(root_conf_path):
            try:
                with open(root_conf_path, "r", encoding="utf-8") as f:
                    root_conf = json.load(f)
                current_algorithm = root_conf.get("algorithm")
            except Exception:
                current_algorithm = None

    if not current_algorithm:
        return []

    if not os.path.isdir(runs_dir):
        return []

    matching_configs = []
    try:
        for item in os.listdir(runs_dir):
            run_path = os.path.join(runs_dir, item)
            if not os.path.isdir(run_path):
                continue

            conf_path = os.path.join(run_path, "conf.json")
            if not os.path.isfile(conf_path):
                candidate = None
                try:
                    for child in os.listdir(run_path):
                        child_conf = os.path.join(run_path, child, "conf.json")
                        if os.path.isfile(child_conf):
                            candidate = child_conf
                            break
                except Exception:
                    pass
                conf_path = candidate if candidate else None

            if conf_path and os.path.isfile(conf_path):
                try:
                    with open(conf_path, "r", encoding="utf-8") as f:
                        conf = json.load(f)
                    if conf.get("algorithm") == current_algorithm:
                        matching_configs.append(conf)
                except Exception:
                    continue
    except Exception:
        return matching_configs

    return matching_configs


def main():
    project_root = os.path.dirname(os.path.abspath(__file__))
    env_path = os.path.join(project_root, ".env")
    runs_root = os.path.join(project_root, "runs_data")

    session_name = get_session_name(env_path)
    output_zip = os.path.join(project_root, f"{session_name}_runs_results.zip")

    num_runs, num_files = create_zip(session_name, runs_root, output_zip)

    # Upload to Telegram
    chat_id, bot_token = get_telegram_credentials(env_path)
    upload_to_telegram(output_zip, chat_id, bot_token)
    print(f"Uploaded zip to Telegram chat {chat_id}")
    print(f"Created zip: {output_zip}")
    print(f"Runs included: {num_runs}")
    print(f"Files included: {num_files}")

def upload_to_telegram(file_path, chat_id, bot_token):
    try:
        import requests  # imported lazily to avoid global dependency
    except ImportError:
        raise RuntimeError("The 'requests' package is required. Install via 'pip install requests'.")

    url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
    with open(file_path, "rb") as f:
        files = {"document": (os.path.basename(file_path), f, "application/zip")}
        data = {"chat_id": str(chat_id)}
        resp = requests.post(url, data=data, files=files, timeout=60)

    if resp.status_code != 200:
        raise RuntimeError(f"Telegram upload failed: {resp.status_code} {resp.text}")


if __name__ == "__main__":
    # GEt environment variables
    load_dotenv()

    try:
        if "--check_duplicate_config" in sys.argv: 
            try:
                conf = json.loads(sys.argv[sys.argv.index("--check_duplicate_config") + 1])
                # Remove the keys that may change and are unnecessary for comparing
                conf.pop("tuning", None)
                conf.pop("tune", None)
                conf.pop("debug", None)
                conf.pop("local_backup", None)
                conf.pop("upload_to_cloud", None)
                conf.pop("train_max_time", None)
                conf.pop("stop_learning_at_win_percent", None)
                conf.pop("max_run_time", None)
                conf.pop("extra_info", None)
                conf.pop("agents", None)
                conf.pop("continue_run", None)
                conf.pop("finished", None)

                # Get all configs that have the same algorithm (located in runs_data directory)
                similarConfigs = getMatchingRunConfigs(runs_dir="runs_data", current_algorithm=conf["algorithm"])

                found_duplicate = False
                for _conf in similarConfigs:
                    # See if training has finished
                    _finished = True if str(_conf.get("finished", None)) == "True" else False if str(_conf.get("finished", None)) == "False" else None

                    # Remove the keys that may change and are unnecessary for comparing
                    _conf.pop("tuning", None)
                    _conf.pop("tune", None)
                    _conf.pop("debug", None)
                    _conf.pop("local_backup", None)
                    _conf.pop("upload_to_cloud", None)
                    _conf.pop("train_max_time", None)
                    _conf.pop("stop_learning_at_win_percent", None)
                    _conf.pop("max_run_time", None)
                    _conf.pop("extra_info", None)
                    _conf.pop("agents", None)
                    _conf.pop("continue_run", None)
                    _conf.pop("finished", None)

                    if conf == _conf and _finished == True:
                        print("true: finished duplicate found")
                        found_duplicate = True
                        break
                    elif conf == _conf and _finished == False:
                        print("false: unfinished duplicate found")
                        found_duplicate = True
                        break

                if not found_duplicate:
                    print("false: no duplicates found")

            except Exception as e:
                raise Exception("Invalid JSON format in --check_duplicate_config parameter. Configuration should be passed exactly after --check_duplicate_config flag.")
        elif "--upload_to_telegram" in sys.argv:
            main()
        elif "--analyze_results" in sys.argv:
            # After --analyze_results, paths to the zip files to be analyze should be passed 
            _start = sys.argv.index("--analyze_results") + 1
            files = sys.argv[_start:]

            if files == []:
                print("No files to analyze. Please pass paths to the zip files to be analyze after --analyze_results flag.")
                sys.exit(1)

            # Get file location
            _loc = os.path.dirname(os.path.abspath(__file__))

            # Refresh the analysis directory
            if os.path.isdir(os.path.join(_loc, "analysis")):
                shutil.rmtree(os.path.join(_loc, "analysis"), ignore_errors=False)
            os.makedirs(os.path.join(_loc, "analysis"), exist_ok=True)

            # Runs dataframe
            data = []

            for file in files:
                sessionName = (os.path.basename(file).split(".")[0]).split("_")[0]
                with zipfile.ZipFile(file, 'r') as zip_ref:
                    os.makedirs(os.path.join(_loc, "analysis", sessionName), exist_ok=True)
                    zip_ref.extractall(os.path.join(_loc, "analysis", sessionName))

                    runsDataDf = pd.read_excel(os.path.join(_loc, "analysis", sessionName, f"{sessionName}_trainingRunsData.xlsx"), header=[0, 1])
                    data += runsDataDf.to_dict(orient="records")

            # Save the dataframe to excel
            with ExcelWriter(os.path.join(_loc, "analysis", "AllRuns.xlsx"), engine="openpyxl") as writer:
                df = pd.DataFrame(data, columns=runsDataDf.columns)
                df.to_excel(writer, sheet_name="Sheet1")

                # Get the workbook and sheet
                workbook = writer.book
                sheet = workbook["Sheet1"]

                for col in sheet.columns:
                    for cell in col:
                        if isinstance(cell.value, str) and "Unnamed" in cell.value:
                            _postMergeValue = sheet.cell(row=cell.row-1, column=cell.column).value
                            sheet.merge_cells(start_row =  cell.row-1, start_column = cell.column, end_row = cell.row, end_column = cell.column)
                            sheet.cell(row=cell.row-1, column=cell.column).value = _postMergeValue

                # Add images columns
                _colEpisodeReward = sheet.max_column + 1
                sheet.insert_cols(idx = _colEpisodeReward)
                sheet.cell(row=1, column= _colEpisodeReward).value = "episodeReward"
                _colTrainingProcess = sheet.max_column + 1
                sheet.insert_cols(idx = _colTrainingProcess)
                sheet.cell(row=1, column= _colTrainingProcess).value = "trainingProcess"
                _colGradientNorms = sheet.max_column + 1
                sheet.insert_cols(idx = _colGradientNorms)
                sheet.cell(row=1, column= _colGradientNorms).value = "gradientNorms"
                
                # Delete empty row
                sheet.delete_rows(idx=3)

                # Add links
                i = 0
                for row in sheet.iter_rows():
                    i+=1
                    if i <2: continue
                    session = row[1].value
                    runNum = row[2].value

                    try:
                        if os.path.exists(os.path.join(_loc, "analysis", session, f"{(runNum)}", "episode_rewards.png")):
                            row[_colEpisodeReward-1].value = "Link"
                            row[_colEpisodeReward-1].hyperlink = f"file:///{os.path.join(_loc, 'analysis', session, f'{(runNum)}', f'episode_rewards.png')}"
                            row[_colEpisodeReward-1].style = "Hyperlink"
                        else: 
                            row[_colEpisodeReward-1].value = "No Data"
                        
                        if os.path.exists(os.path.join(_loc, "analysis", session, f"{(runNum)}", "training_process.png")):
                            row[_colTrainingProcess-1].value = "Link"
                            row[_colTrainingProcess-1].hyperlink = f"file:///{os.path.join(_loc, 'analysis', session, f'{(runNum)}', f'training_process.png')}"
                            row[_colTrainingProcess-1].style = "Hyperlink"
                        else: 
                            row[_colTrainingProcess-1].value = "No Data"    
                        
                        if os.path.exists(os.path.join(_loc, "analysis", session, f"{(runNum)}", "gradient_norms.png")):
                            row[_colGradientNorms-1].value = "Link"
                            row[_colGradientNorms-1].hyperlink = f"file:///{os.path.join(_loc, 'analysis', session, f'{(runNum)}', f'gradient_norms.png')}"
                            row[_colGradientNorms-1].style = "Hyperlink"
                        else:
                            row[_colGradientNorms-1].value = "No Data"    

                    except Exception as e:
                        print(f"Error -> {i} - {e}")
                        traceback.print_exc() 
                        continue
                    

                # Style cells
                for row in sheet.iter_rows():
                    for cell in row:
                        cell.font = Font(name="Calibri", bold=False)
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                # Set cell widths
                for col in sheet.columns:
                    max_length = 0
                    col_letter = get_column_letter(col[0].column)

                    for cell in col:
                        try:
                            cell_len = len(str(cell.value))
                            if cell_len > max_length:
                                max_length = cell_len
                        except:
                            pass

                    # Add small extra space for padding
                    sheet.column_dimensions[col_letter].width = max_length + .2
                    

        else:
            print("No valid flag found. Acceptable flags are --check_duplicate_config , --upload_to_telegram and --analyze_results")
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)